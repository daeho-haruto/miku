import discord
import asyncio
import urllib.parse
import urllib
import youtube_dl
import openpyxl
import datetime
import time
from discord.utils import get
from discord.ext import commands
import os
from os.path import join, dirname
from dotenv import load_dotenv

dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

token = os.environ.get("MIKU_PSW")


client = commands.Bot(command_prefix='!')
# intents = discord.Intents.default()
# client = discord.Client(intents=intents)

@client.event
async def on_ready():
    print("봇 준비 완료!")
    print(client.user.name)
    print("---------------")
    game = discord.Game("!명령어 라고 쳐주세요")
    await client.change_presence(status=discord.Status.online, activity=game)




@client.command()
async def 미쿠(ctx):
    await ctx.send("初音ミク 登場!")

@client.command()
async def 명령어(ctx):
    embed = discord.Embed(title="명령어", color=0x62c1cc)
    embed.set_thumbnail(url="https://i.imgur.com/094U5yG.jpg")
    embed.add_field(name="!미쿠",value="`하츠네 미쿠 등장!`")
    embed.add_field(name="!따라해",value="`따라하기`",inline=False)
    embed.add_field(name="!입장",value="`음성 채팅 입장`")
    embed.add_field(name="!재생",value="`노래 재생 (재생 뒤 youtube url)`")
    embed.add_field(name="!퇴장",value="`음성 채팅 퇴장`",inline=False)
    embed.add_field(name="!경고",value="`경고주기`")
    embed.add_field(name="!경고보기",value="`경고보기`")
    embed.add_field(name="!경고초기화",value="`경고초기화`",inline=False)
    embed.add_field(name="!기능추가",value="`기능추가`")
    embed.add_field(name="!기능보기",value="`기능보기`")
    embed.add_field(name="!기능삭제",value="`기능삭제`")
    embed.add_field(name="!기능초기화",value="`기능초기화`",inline=False)
    embed.add_field(name="!타이머",value="`타이머(초)`",inline=False)
    embed.add_field(name="!알람추가",value="`알람추가`")
    embed.add_field(name="!알람보기",value="`알람보기`")
    embed.add_field(name="!알람삭제",value="`알람삭제`")
    embed.add_field(name="!알람초기화",value="`알람초기화`")
    embed.add_field(name="!투표",value="`ex)!투표 투표제목/(투표1)/(투표2)`")
    embed.add_field(name="!신음",value="`미쿠의 신음소리`",inline=False)
    embed.add_field(name="!메세지삭제",value="`메세지 삭제 (개수)`",inline=False)
    embed.add_field(name="!사진목록1",value="`목록 1`")
    embed.add_field(name="!사진목록2",value="`목록 2`")
    embed.add_field(name="!사진",value="`사진(숫자)`")

    await ctx.send(embed=embed)

@client.command()
async def 따라해(ctx):
    msg = ctx.message.content[5:]
    await ctx.send(msg)

@client.command()
async def 입장(ctx):
    await ctx.message.author.voice.channel.connect()
    await ctx.message.channel.send("보이스채널 입장합니다.")
    
@client.command()
async def 퇴장(ctx):
    global voice
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc
        
    await voice.disconnect()
    await ctx.message.channel.send("보이스채널 퇴장합니다.")

@client.command()
async def 재생(ctx):
    global voice
    for vc in client.voice_clients:
        if vc.guild == ctx.message.guild:
            voice = vc

    url = ctx.message.content.split(" ")[1]
    option = {
        # 'outtmpl' : "file/" + url.split('=')[1] + ".mp3"
        'outtmpl' : "file/" + url.split('=')[1] + ".mp3",
        'format': 'bestaudio/best',
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }]
    }

    with youtube_dl.YoutubeDL(option) as ydl:
        ydl.download([url])
        info = ydl.extract_info(url, download=False)
        title = info["title"]

    voice.play(discord.FFmpegPCMAudio("file/" + url.split('=')[1] + ".mp3"))
    await ctx.message.channel.send(title + "을 재생합니다.")

@client.command()
async def 경고(ctx):
    msg = ctx.message.content[4:]
    file = openpyxl.load_workbook("경고.xlsx")
    sheet = file.active
    i = 1
    while True:
        if sheet["A" + str(i)].value == str(msg):
            sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
            file.save("경고.xlsx")
            await ctx.message.channel.send("경고를 1회 받았습니다.")
            if sheet["B" + str(i)].value == 2:
                await ctx.message.channel.send("누적 2회")
            elif sheet["B" + str(i)].value == 3:
                await ctx.message.channel.send("누적 3회")
            elif sheet["B" + str(i)].value == 4:
                await ctx.message.channel.send("누적 4회")
            elif sheet["B" + str(i)].value == 5:
                await ctx.message.channel.send("누적 5회")
            elif sheet["B" + str(i)].value == 6:
                await ctx.message.channel.send("누적 6회")
            elif sheet["B" + str(i)].value == 7:
                await ctx.message.channel.send("누적 7회")
            elif sheet["B" + str(i)].value == 8:
                await ctx.message.channel.send("누적 8회")
            elif sheet["B" + str(i)].value == 9:
                await ctx.message.channel.send("누적 9회")
            elif sheet["B" + str(i)].value == 10:
                await ctx.message.channel.send("누적 10회 퇴장조치")
            break
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(msg)
            sheet["B" + str(i)].value = 1
            file.save("경고.xlsx")
            await ctx.message.channel.send("경고를 1회 받았습니다.")
            await ctx.message.channel.send("누적 1회")
            break
        i+=1

@client.command()
async def 경고보기(ctx):
    file = openpyxl.load_workbook("경고.xlsx")
    sheet = file.active
    i=1
    try:
        for i in range(1,100):
            await ctx.message.channel.send(sheet.cell(i,1).value)
            await ctx.message.channel.send(sheet.cell(i,2).value)
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 경고초기화(ctx):
    file = openpyxl.load_workbook("경고.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    sheet.delete_cols(1)
    file.save("경고.xlsx")
    await ctx.message.channel.send("경고가 초기화 되었습니다.")



@client.command()
async def 기능추가(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            sheet["A" + str(i)].value = str(msg)
            file.save("기능추가.xlsx")
            await ctx.message.channel.send(msg + " 기능이 추가 되었습니다.")
            break
        i+=1

@client.command()
async def 기능보기(ctx):
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet = file.active
    i=1
    try:
        for i in range(1,100):
            await ctx.message.channel.send(sheet.cell(i,1).value)
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 기능삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == str(msg):
            sheet.delete_rows(i)
            file.save("기능추가.xlsx")
            await ctx.message.channel.send(msg + " 기능이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 기능초기화(ctx):
    file = openpyxl.load_workbook("기능추가.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("기능추가.xlsx")
    await ctx.message.channel.send("기능이 초기화 되었습니다.")

@client.command()
async def 타이머(ctx):
    msg = ctx.message.content[5:]
    sec = int(msg)

    for i in range(sec, 0, -1):
        time.sleep(1)
        await ctx.message.channel.send(embed=discord.Embed(description= str(i) + "초"))
    
    await ctx.message.channel.send(embed=discord.Embed(description='타이머 종료'))

@client.command()
async def 알람추가(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        if sheet["A" + str(i)].value == None:
            li = msg.split(" ")
            da1 = li.pop(0)
            da2 = li.pop(0)
            da3 = li.pop(0)
            da4 = li.pop(0)
            dd = da1 + da2 + da3 + da4
            sheet["A" + str(i)].value = str(dd)
            file.save("알람.xlsx")
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 추가되었습니다.")
            break
        i+=1

@client.command()
async def 알람삭제(ctx):
    msg = ctx.message.content[6:]
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    while True:
        li = msg.split("/")
        da1 = li.pop(0)
        da2 = li.pop(0)
        da3 = li.pop(0)
        da4 = li.pop(0)
        dd = da1 + da2 + da3 + da4 
        if sheet["A" + str(i)].value == str(dd):
            sheet.delete_rows(i)
            file.save("알람.xlsx")
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분 알람이 삭제 되었습니다.")
            break
        i+=1

@client.command()
async def 알람보기(ctx):
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    i=1
    try:
        for i in range(1,100):
            dd = sheet.cell(i,1).value
            month = dd[:2]
            day = dd[2:4]
            hour = dd[4:6]
            minu = dd[6:]
            await ctx.message.channel.send(month + '월 ' + day + "일 " + hour + '시 ' + minu + "분")
            i+=1
    except :
        await ctx.message.channel.send("--------")
        await ctx.message.channel.send("여기까지")

@client.command()
async def 알람초기화(ctx):
    file = openpyxl.load_workbook("알람.xlsx")
    sheet  = file.active
    sheet.delete_cols(1)
    file.save("알람.xlsx")
    await ctx.message.channel.send("알람이 초기화 되었습니다.")
    
# @client.commad()
# async def 회의알람():
#     file = openpyxl.load_workbook("알람.xlsx")
#     sheet  = file.active
#     i = 1
#     while True:
#         dt = datetime.datetime.now()
#         dt = dt.strftime('%Y/%m/%d %H:%M')

#         li = dt 
#         li1 = li.split(" ")

#         day = li1.pop(0)
#         time = li1.pop(0)

#         day = day.split("/")
#         time = time.split(":")

#         day1 = day.pop(0)
#         time1 = time.pop(0)

#         day2 = day.pop(0)
#         time2 = time.pop(0)

#         day3 = day.pop(0)

#         last = day1 + day2 + day3 + time1 + time2
        
#         print(last)

#         if sheet["A" + str(i)].value == last:
#             sheet.delete_rows(i)
#             file.save("알람.xlsx")
#             await ctx.message.channel.send("@averyone")
#             break
#         i+=1

@client.command()
async def 삭제(ctx):
    number = int(ctx.message.content.split()[1])
    await ctx.message.channel.purge(limit=number + 1)
    await ctx.message.channel.send(f"{number}개 메세지 삭제완료")

@client.command()
async def 투표(ctx):
    vote = ctx.message.content[4:].split("/")
    await ctx.message.channel.send("★투표 - " + vote[0])
    for i in range(1, len(vote)):
        choose = await ctx.message.channel.send("```" + vote[i] + "```")
        await choose.add_reaction('👍')

# @client.command()
# async def 공건불러(ctx):
#     count = ctx.message.content[6:]
#     for i in range(0,int(count)):
#         msg = "<@{}>".format(str(354168092643033090))
#         time.sleep(1)
#         await ctx.message.channel.send(msg)

# @client.command()
# async def 주기알람(ctx):
#     msg = ctx.message.content[6:]
#     await ctx.message.channel.send(msg + "일 마다 알람이 울립니다.")
#     await ctx.message.channel.send("*단 프로그램이 켜져있어야 합니다. (또는 호스팅)*")
#     msg = int(msg)
#     while True:
#         time.sleep(86400*msg)
#         await ctx.message.channel.send("주기 알람 @everyone")

# @client.command()
# async def 주기알람(ctx):
#     msg = ctx.message.content[6:]
#     await ctx.message.channel.send(msg + "시간 마다 알람이 울립니다.")
#     msg = int(msg)
#     while True:
#         time.sleep(3600*msg)
#         await ctx.message.channel.send("주기 알람")

@client.command()
async def 사진목록1(ctx):
    embed = discord.Embed(title="사진", color=0x62c1cc)
    embed.add_field(name="1", value="최초미쿠")
    embed.add_field(name="2", value="보컬미쿠")
    embed.add_field(name="3", value="사쿠라미쿠")
    embed.add_field(name="4", value="하츄네미쿠")
    embed.add_field(name="5", value="자츠네미쿠")
    embed.add_field(name="6", value="하츠네즈미")
    embed.add_field(name="7", value="야미네아쿠")
    embed.add_field(name="8", value="노로네키쿠")
    embed.add_field(name="9", value="카루네시에")
    embed.add_field(name="10", value="유키미쿠")
    embed.add_field(name="11", value="레이싱미쿠")
    embed.add_field(name="12", value="보틀미쿠")
    embed.add_field(name="13", value="아니마사식미쿠")
    embed.add_field(name="14", value="Lat식미쿠")
    embed.add_field(name="15", value="나노하1052식미쿠")
    embed.add_field(name="16", value="러브식미쿠")
    embed.add_field(name="17", value="온다식미쿠")
    embed.add_field(name="18", value="어피어런스미쿠")
    embed.add_field(name="19", value="xs식미쿠")
    embed.add_field(name="20", value="Tda식미쿠")
    await ctx.send(embed=embed)
@client.command()
async def 사진목록2(ctx):
    embed = discord.Embed(title="사진", color=0x62c1cc)
    embed.add_field(name="21", value="REMmaple식V3미쿠")
    embed.add_field(name="22", value="키오식미쿠")
    embed.add_field(name="23", value="디바풍미쿠")
    embed.add_field(name="24", value="아피미쿠")
    embed.add_field(name="25", value="Ula식미쿠")
    embed.add_field(name="26", value="마레욘식미쿠")
    embed.add_field(name="27", value="YYB식미쿠")
    embed.add_field(name="28", value="Gency식미쿠")
    embed.add_field(name="29", value="시온식미쿠")
    embed.add_field(name="30", value="하츠키식미쿠")
    embed.add_field(name="31", value="별이빛나는미쿠")
    embed.add_field(name="32", value="피쿠치식미쿠")
    embed.add_field(name="33", value="유키하네식미쿠")
    embed.add_field(name="34", value="코론식미쿠")
    embed.add_field(name="35", value="로쿠고식미쿠")
    embed.add_field(name="36", value="Digitrevx식미쿠")
    embed.add_field(name="37", value="미야식미쿠")
    embed.add_field(name="38", value="넨드로이드미쿠")
    embed.add_field(name="39", value="고양이미쿠")
    embed.add_field(name="40", value="비키니미쿠")
    await ctx.send(embed=embed)

@client.command()
async def 사진(ctx):
    embed = discord.Embed(imestamp=ctx.message.created_at,color=0x62c1cc)
    msg = ctx.message.content[4:]
    msg = int(msg)
    if msg == 1:
        embed.set_image(url="https://i.imgur.com/JG7JXpA.jpg")
        
        await ctx.message.channel.send(embed=embed)
    elif msg == 2:
        embed.set_image(url="https://i.imgur.com/tWvu4xk.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 3:
        embed.set_image(url="https://i.imgur.com/lmTycvB.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 4:
        embed.set_image(url="https://i.imgur.com/khoYN5r.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 5:
        embed.set_image(url="https://i.imgur.com/9gq8k25.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 6:
        embed.set_image(url="https://i.imgur.com/vSmGhGW.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 7:
        embed.set_image(url="https://i.imgur.com/1UJev0d.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 8:
        embed.set_image(url="https://i.imgur.com/2vfOEKZ.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 9:
        embed.set_image(url="https://i.imgur.com/vtDEhWB.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 10:
        embed.set_image(url="https://i.imgur.com/zIyeGTa.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 11:
        embed.set_image(url="https://i.imgur.com/GBKovkx.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 12:
        embed.set_image(url="https://i.imgur.com/OIoGGct.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 13:
        embed.set_image(url="https://i.imgur.com/QS6yFgZ.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 14:
        embed.set_image(url="https://i.imgur.com/Z07ta5J.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 15:
        embed.set_image(url="https://i.imgur.com/WB8LpT5.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 16:
        embed.set_image(url="https://i.imgur.com/mKwwOnc.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 17:
        embed.set_image(url="https://i.imgur.com/0QZCgMP.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 18:
        embed.set_image(url="https://i.imgur.com/S2Rgjlo.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 19:
        embed.set_image(url="https://i.imgur.com/yzkDGqk.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 20:
        embed.set_image(url="https://i.imgur.com/yupLbnj.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 21:
        embed.set_image(url="https://i.imgur.com/G1R1vZd.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 22:
        embed.set_image(url="https://i.imgur.com/mamFxFC.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 23:
        embed.set_image(url="https://i.imgur.com/NX0dJvE.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 24:
        embed.set_image(url="https://i.imgur.com/OuPJqcx.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 25:
        embed.set_image(url="https://i.imgur.com/1jUDYK8.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 26:
        embed.set_image(url="https://i.imgur.com/F3TtqoP.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 27:
        embed.set_image(url="https://i.imgur.com/spNANjO.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 28:
        embed.set_image(url="https://i.imgur.com/bPiLOaz.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 29:
        embed.set_image(url="https://i.imgur.com/6ehrsMQ.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 30:
        embed.set_image(url="https://i.imgur.com/1YBuo7g.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 31:
        embed.set_image(url="https://i.imgur.com/vJvTgPs.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 32:
        embed.set_image(url="https://i.imgur.com/5gQayVl.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 33:
        embed.set_image(url="https://i.imgur.com/Z1Q73LN.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 34:
        embed.set_image(url="https://i.imgur.com/RHsQmiN.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 35:
        embed.set_image(url="https://i.imgur.com/sfIcE2R.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 36:
        embed.set_image(url="https://i.imgur.com/0YK0ILZ.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 37:
        embed.set_image(url="https://i.imgur.com/gIMSv9x.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 38:
        embed.set_image(url="https://i.imgur.com/xd1xwiT.png")
        await ctx.message.channel.send(embed=embed)
    elif msg == 39:
        embed.set_image(url="https://i.imgur.com/DpZHG2D.jpg")
        await ctx.message.channel.send(embed=embed)
    elif msg == 40:
        embed.set_image(url="https://i.imgur.com/OhpQosJ.jpg")
        await ctx.message.channel.send(embed=embed)
    else:
        msg = str(msg)
        await ctx.message.channel.send(msg + " 번호는 없습니다")
        embed.set_image(url="https://i.imgur.com/cmSfQey.gif")
        await ctx.message.channel.send(embed=embed)



client.run(token)
